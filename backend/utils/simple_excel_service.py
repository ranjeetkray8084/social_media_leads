"""
Simplified Excel Service without pandas dependency
"""

import os
import logging
from datetime import datetime
from typing import List, Dict
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class SimpleExcelService:
    def __init__(self):
        """Initialize Excel service"""
        self.output_dir = '../data/exports'
        os.makedirs(self.output_dir, exist_ok=True)
        logging.info("Simple Excel Service initialized")
    
    def export_leads_to_excel(self, leads: List[Dict], filename: str = None) -> Dict:
        """Export leads to Excel file"""
        try:
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"leads_export_{timestamp}.xlsx"
            
            filepath = os.path.join(self.output_dir, filename)
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Leads"
            
            # Headers
            headers = [
                "Name", "Username", "Phone", "Email", "WhatsApp", "Social Handle",
                "Requirement", "Budget", "Location", "Source", "Lead Score",
                "Buying Intent", "Timeline", "Contact Method", "Post URL", "Timestamp"
            ]
            
            # Style headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Add headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Add data
            for row, lead in enumerate(leads, 2):
                ws.cell(row=row, column=1, value=lead.get('name', ''))
                ws.cell(row=row, column=2, value=lead.get('username', ''))
                ws.cell(row=row, column=3, value=lead.get('phone', ''))
                ws.cell(row=row, column=4, value=lead.get('email', ''))
                ws.cell(row=row, column=5, value=lead.get('whatsapp', ''))
                ws.cell(row=row, column=6, value=lead.get('social_handle', ''))
                ws.cell(row=row, column=7, value=lead.get('requirement', ''))
                ws.cell(row=row, column=8, value=lead.get('budget', ''))
                ws.cell(row=row, column=9, value=lead.get('location', ''))
                ws.cell(row=row, column=10, value=lead.get('source', ''))
                ws.cell(row=row, column=11, value=lead.get('lead_score', ''))
                ws.cell(row=row, column=12, value=lead.get('buying_intent', ''))
                ws.cell(row=row, column=13, value=lead.get('timeline', ''))
                ws.cell(row=row, column=14, value=lead.get('contact_method', ''))
                ws.cell(row=row, column=15, value=lead.get('post_url', ''))
                ws.cell(row=row, column=16, value=lead.get('timestamp', ''))
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save file
            wb.save(filepath)
            
            logging.info(f"Excel file created: {filepath}")
            
            return {
                'success': True,
                'filename': filename,
                'filepath': filepath,
                'download_url': f'/api/download/{filename}',
                'leads_count': len(leads)
            }
            
        except Exception as e:
            logging.error(f"Error creating Excel file: {e}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def get_export_history(self) -> List[Dict]:
        """Get list of exported files"""
        try:
            files = []
            if os.path.exists(self.output_dir):
                for filename in os.listdir(self.output_dir):
                    if filename.endswith('.xlsx'):
                        filepath = os.path.join(self.output_dir, filename)
                        stat = os.stat(filepath)
                        files.append({
                            'filename': filename,
                            'size': stat.st_size,
                            'created': datetime.fromtimestamp(stat.st_ctime).isoformat(),
                            'download_url': f'/api/download/{filename}'
                        })
            
            # Sort by creation time (newest first)
            files.sort(key=lambda x: x['created'], reverse=True)
            return files
            
        except Exception as e:
            logging.error(f"Error getting export history: {e}")
            return []
    
    def delete_export_file(self, filename: str) -> bool:
        """Delete an exported file"""
        try:
            filepath = os.path.join(self.output_dir, filename)
            if os.path.exists(filepath):
                os.remove(filepath)
                logging.info(f"Deleted file: {filename}")
                return True
            return False
            
        except Exception as e:
            logging.error(f"Error deleting file {filename}: {e}")
            return False
