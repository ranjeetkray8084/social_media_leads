"""
Excel Service for Lead Export
Handles Excel file creation and lead data export
"""

import pandas as pd
import os
import logging
from datetime import datetime
from typing import List, Dict
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class ExcelService:
    def __init__(self):
        """Initialize Excel service"""
        self.output_dir = '../data/exports'
        os.makedirs(self.output_dir, exist_ok=True)
        logging.info("Excel Service initialized")
    
    def export_leads_to_excel(self, leads: List[Dict], filename: str = None) -> str:
        """Export leads to Excel file"""
        try:
            if not leads:
                logging.warning("No leads to export")
                return None
            
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"social_media_leads_{timestamp}.xlsx"
            
            # Ensure .xlsx extension
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
            
            filepath = os.path.join(self.output_dir, filename)
            
            # Prepare data for Excel
            excel_data = []
            for lead in leads:
                excel_row = {
                    'Name': lead.get('name', 'Unknown'),
                    'Phone': lead.get('phone', ''),
                    'Email': lead.get('email', ''),
                    'Requirement': lead.get('requirement', ''),
                    'Location': lead.get('location', ''),
                    'Budget': lead.get('budget', ''),
                    'Source': lead.get('source', ''),
                    'Lead Score': lead.get('lead_score', 0),
                    'Username': lead.get('username', ''),
                    'Language': lead.get('language', 'English'),
                    'Confidence': lead.get('confidence', 0),
                    'Original Content': lead.get('original_content', ''),
                    'Social Media URL': lead.get('social_media_url', ''),
                    'Extracted At': lead.get('extracted_at', ''),
                    'Status': lead.get('status', 'NEW'),
                    'Action': lead.get('action', 'CONTACT')
                }
                excel_data.append(excel_row)
            
            # Create DataFrame
            df = pd.DataFrame(excel_data)
            
            # Create Excel writer with formatting
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Social Media Leads', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Social Media Leads']
                
                # Apply formatting
                self._format_excel_sheet(worksheet, workbook)
            
            logging.info(f"Exported {len(leads)} leads to {filepath}")
            return filepath
            
        except Exception as e:
            logging.error(f"Error exporting leads to Excel: {e}")
            return None
    
    def _format_excel_sheet(self, worksheet, workbook):
        """Apply formatting to Excel sheet"""
        try:
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Format headers
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Max width 50
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Format data cells
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Color code lead scores
                    if cell.column == 8 and cell.value:  # Lead Score column
                        try:
                            score = int(cell.value)
                            if score >= 8:
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
                            elif score >= 6:
                                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
                            else:
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
                        except:
                            pass
            
        except Exception as e:
            logging.error(f"Error formatting Excel sheet: {e}")
    
    def create_summary_sheet(self, leads: List[Dict], filepath: str):
        """Create a summary sheet with statistics"""
        try:
            from openpyxl import load_workbook
            
            # Load existing workbook
            workbook = load_workbook(filepath)
            
            # Create summary data
            total_leads = len(leads)
            platforms = {}
            locations = {}
            lead_scores = []
            
            for lead in leads:
                # Platform statistics
                platform = lead.get('source', 'Unknown')
                platforms[platform] = platforms.get(platform, 0) + 1
                
                # Location statistics
                location = lead.get('location', 'Not specified')
                locations[location] = locations.get(location, 0) + 1
                
                # Lead scores
                score = lead.get('lead_score', 0)
                if score:
                    lead_scores.append(score)
            
            # Create summary data
            summary_data = [
                ['Summary Report', ''],
                ['Generated At', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['Total Leads', total_leads],
                ['', ''],
                ['Platform Distribution', ''],
            ]
            
            for platform, count in platforms.items():
                summary_data.append([platform, count])
            
            summary_data.extend([
                ['', ''],
                ['Location Distribution', ''],
            ])
            
            for location, count in list(locations.items())[:10]:  # Top 10 locations
                summary_data.append([location, count])
            
            if lead_scores:
                summary_data.extend([
                    ['', ''],
                    ['Lead Score Statistics', ''],
                    ['Average Score', round(sum(lead_scores) / len(lead_scores), 2)],
                    ['Highest Score', max(lead_scores)],
                    ['Lowest Score', min(lead_scores)],
                    ['High Quality Leads (8+)', len([s for s in lead_scores if s >= 8])],
                    ['Medium Quality Leads (6-7)', len([s for s in lead_scores if 6 <= s < 8])],
                    ['Low Quality Leads (<6)', len([s for s in lead_scores if s < 6])],
                ])
            
            # Create summary sheet
            summary_sheet = workbook.create_sheet('Summary')
            
            for row_idx, row_data in enumerate(summary_data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = summary_sheet.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Format headers
                    if row_idx == 1 or value.endswith('Distribution') or value.endswith('Statistics'):
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            # Save workbook
            workbook.save(filepath)
            logging.info("Summary sheet created successfully")
            
        except Exception as e:
            logging.error(f"Error creating summary sheet: {e}")
    
    def get_export_history(self) -> List[Dict]:
        """Get list of exported files"""
        try:
            files = []
            if os.path.exists(self.output_dir):
                for filename in os.listdir(self.output_dir):
                    if filename.endswith('.xlsx'):
                        filepath = os.path.join(self.output_dir, filename)
                        stat = os.stat(filepath)
                        files.append({
                            'filename': filename,
                            'filepath': filepath,
                            'size': stat.st_size,
                            'created': datetime.fromtimestamp(stat.st_ctime).isoformat(),
                            'modified': datetime.fromtimestamp(stat.st_mtime).isoformat()
                        })
            
            # Sort by creation time (newest first)
            files.sort(key=lambda x: x['created'], reverse=True)
            return files
            
        except Exception as e:
            logging.error(f"Error getting export history: {e}")
            return []
    
    def delete_export_file(self, filename: str) -> bool:
        """Delete an exported file"""
        try:
            filepath = os.path.join(self.output_dir, filename)
            if os.path.exists(filepath):
                os.remove(filepath)
                logging.info(f"Deleted export file: {filename}")
                return True
            return False
        except Exception as e:
            logging.error(f"Error deleting export file {filename}: {e}")
            return False
